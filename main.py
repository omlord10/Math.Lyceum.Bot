from openpyxl import load_workbook
import mems
import telebot
import random
from telebot import types
from random import randint
import sqlite3
import xlrd
import os
from dotenv import load_dotenv, find_dotenv


load_dotenv(find_dotenv())

conn = sqlite3.connect('users.db', check_same_thread=False)
cursor = conn.cursor()
cursor.execute('''
CREATE TABLE IF NOT EXISTS tickets (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    user_id INTEGER NOT NULL,
    problem TEXT NOT NULL,
    ticket_id,
    status TEXT DEFAULT 'Open'
)
''')

bot = telebot.TeleBot(os.getenv('OLD_TOKEN'))

standart_list = []
all_username_list = []
ban_list = []
moder_list = []
admin_list = []
names_standart_list = []
memory = []
memory2 = []


cursor.execute("SELECT tg_id FROM students WHERE role='standart'")
results = cursor.fetchall()
for row in results:
    standart_list.append(row[0])

cursor.execute('SELECT name, tg_id, user_class, role FROM students')
results = cursor.fetchall()
users_dict = {}
for row in results:
    name = row[0]
    tg_id = row[1]
    user_class = row[2]
    role = row[3]
    users_dict[name] = {'name': name, 'tg_id': tg_id, 'user_class': user_class, 'role': role}

cursor.execute("SELECT tg_id FROM students WHERE role='admin'")
results = cursor.fetchall()
for row in results:
    admin_list.append(row[0])

cursor.execute("SELECT tg_id FROM students WHERE role='moder'")
results = cursor.fetchall()
for row in results:
    moder_list.append(row[0])

cursor.execute("SELECT tg_id FROM students WHERE role='banned'")
results = cursor.fetchall()
for row in results:
    ban_list.append(row[0])

cursor.execute("SELECT name FROM students WHERE role='standart'")
results = cursor.fetchall()
for row in results:
    names_standart_list.append(row[0])

@bot.message_handler(commands=['start', 'Start', 'старт', 'Старт'])
def start(message):
    if message.chat.id not in ban_list:
        if message.from_user.last_name != 'None' or message.from_user.last_name is not None:
            mess = f'Привет, <b>{message.from_user.first_name} {message.from_user.last_name}</b>. Я бот ' \
                   f'Математического лицея. Создан помочь всем в школе. ' \
                   f'Нажми /help, чтобы узнать, что я могу.' + '\n' + 'Для получения новостей необходимо указать себя, написав /add_user'
            bot.reply_to(message, mess, parse_mode='html')
        else:
            mess = f'Привет, <b>{message.from_user.first_name}</b>. Я бот Математического лицея. Создан помочь всем в школе. ' \
                   f'Нажми /help, чтобы узнать, что я могу.' + '\n' + 'Для получения новостей необходимо указать себя, написав /add_user'
            bot.send_message(message.chat.id, mess)
    else:
        bot.send_message(message.chat.id, 'BANNED')


@bot.message_handler(commands=['help','Help'])
def help(message):
    if message.chat.id not in ban_list:
        try:
            if message.chat.id not in admin_list and message.chat.id not in moder_list:
                mess = f'Команды можно быстро скопировать, долго нажимая на них.\n' \
                       f'Список всех команд:\n' \
                       f'/start — стартовое сообщение\n' \
                       f'\n' \
                       f'/help — список всех команд\n' \
                       f'\n' \
                       f'/menu  — основное меню\n' \
                       f'\n' \
                       f'/add_user — добавляет пользователя в базу данных. Класс писать в данном формате 1А\n' \
                       f'\n' \
                       f'/classes — меню классов\n' \
                       f'\n' \
                       f'/Id — Id пользователя\n' \
                       f'\n' \
                       f'/Site — школьные соц.сети\n' \
                       f'\n' \
                       f'/Mem — меню мемов\n' \
                       f'\n' \
                       f'/add_ticket — отправляет запрос в тех-поддержку\n' \
                       f'\n' \
                       f'/set_ticket_status ticket_id <pre>Close</pre> или <pre>Open</pre> — изменить статус на открыт или закрыт\n' \
                       f'\n' \
                       f'<pre>/Изменение_в_расписании</pre> — пользователю отправится изменение в расписании'
            elif message.chat.id in admin_list:
                mess = f'Команды можно быстро скопировать, долго нажимая на них.\n' \
                       f'\n' \
                       f'Список всех команд для администратора:\n' \
                       f'\n' \
                       f'/start — стартовое сообщение\n' \
                       f'\n' \
                       f'/help — список всех команд\n' \
                       f'\n' \
                       f'/menu  — основное меню\n' \
                       f'\n' \
                       f'/add_user — добавляет пользователя в базу данных. Класс писать в данном формате 1А\n' \
                       f'\n' \
                       f'/classes — меню классов\n' \
                       f'\n' \
                       f'/Id — Id пользователя\n' \
                       f'\n' \
                       f'/Site — школьные соц.сети\n' \
                       f'\n' \
                       f'/Mem — меню мемов\n' \
                       f'\n' \
                       f'/mem_photo — мем-картинка\n' \
                       f'\n' \
                       f'/mem_text — мем-текст\n' \
                       f'\n' \
                       f'/logs — логи\n' \
                       f'\n' \
                       f'/change_mem — изменяет текущий мем\n' \
                       f'\n' \
                       f'/add_ticket — отправляет запрос в тех-поддержку\n' \
                       f'\n' \
                       f'/tickets — отображает запросы пользователя\n' \
                       f'\n' \
                       f'/set_ticket_status ticket_id <pre>Close</pre> или <pre>Open</pre> — изменить статус на открыт или закрыт\n' \
                       f'\n' \
                       f'При отправке фотографии она будет сохранена на сервере для изменений в расписании. Просьба удалить все фотографии в папке, так как команда /Изменение_в_расписании выдаёт все фотографии в папке.\n' \
                       f'\n' \
                       f'<pre>/Изменение_в_расписании</pre> — пользователю отправится изменение в расписании\n' \
                       f'\n' \
                       f'/delete_photoes путь — удаляет фотографии в папке. Путь имеет вид\n\n<pre>temp/изменение_в_расписании</pre>\n\nили\n\n<pre>temp/mems/photoes</pre>'
            else:
                mess = f'Команды можно быстро скопировать, долго нажимая на них.\n' \
                       f'Список всех команд для модератора:\n' \
                       f'/start — стартовое сообщение\n' \
                       f'\n' \
                       f'/help — список всех команд\n' \
                       f'\n' \
                       f'/menu  — основное меню\n' \
                       f'\n' \
                       f'/add_user — добавляет пользователя в базу данных. Класс писать в данном формате 1А\n' \
                       f'\n' \
                       f'/classes — меню классов\n' \
                       f'\n' \
                       f'/Id — Id пользователя\n' \
                       f'\n' \
                       f'/Site — школьные соц.сети\n' \
                       f'\n' \
                       f'/Mem — меню мемов\n' \
                       f'\n' \
                       f'/mem_photo — мем-картинка\n' \
                       f'\n' \
                       f'/add_ticket — отправляет запрос в тех-поддержку\n' \
                       f'\n' \
                       f'/set_ticket_status ticket_id <pre>Close</pre> или <pre>Open</pre> — изменить статус на открыт или закрыт\n' \
                       f'\n' \
                       f'При отправке фотографии она будет сохранена на сервере для изменений в расписании.\n' \
                       f'\n' \
                       f'<pre>/Изменение_в_расписании</pre> — пользователю отправится изменение в расписании\n'
            bot.send_message(message.chat.id, mess, parse_mode='html')
        except (IndexError, ValueError):
            bot.send_message(message.chat.id,'Ошибка: некорректный ввод')
    else:
        bot.send_message(message.chat.id,'BANNED')

# @bot.message_handler(commands=['change_mem'])
# def change_mem(message):
#     if message.chat.id in moder_list:
#         try:
#             bot.send_message(message.chat.id, 'Успешно! Текст изменён')
#             change_text = message.text[12:]
#             file = open('mems.txt', 'w', encoding='windows-1251')
#             file.write(change_text)
#         except:
#             bot.send_message(message.chat.id,'Ошибка: некорректный ввод')
#     else:
#         bot.send_message(message.chat.id, "You aren't the developer! Only developers can make changes")


@bot.message_handler(commands=['delete_user'])
def delete_user(message):
    if message.chat.id in admin_list:
        bot.send_message(message.chat.id,'Отправь мне id пользователя')
        bot.register_next_step_handler(message, get_delete_id)
    else:
        bot.send_message(message.chat.id,'Нет прав доступа')

def get_delete_id(message):
    try:
        user_id = int(message.text)

        cursor.execute("DELETE FROM students WHERE tg_id=?", (user_id,))
        conn.commit()

        bot.reply_to(message, f"Пользователь с id {user_id} успешно удален.")

    except (IndexError, ValueError):
        bot.reply_to(message, "Ошибка: некорректный id пользователя.")

@bot.message_handler(commands=['delete_photoes'])
def delete_photoes(message):
    if message.chat.id in admin_list:
        try:
            path = message.text.split(' ')[1]
            for file_name in os.listdir(path):
                os.remove(os.path.join(path, file_name))
            bot.reply_to(message, "Фотографии успешно удалены!")
        except IndexError:
            bot.reply_to(message, "Вы не указали путь к директории!")
    else:
        bot.reply_to(message, "У вас нет прав на выполнение данной команды!")


@bot.message_handler(content_types=['photo'])
def handle_admin_photo(message):
    if message.from_user.id in moder_list:
        photo_folder = "temp/mems/photoes"
        saved_photos = []
        photo = message.photo[-1]
        file_id = str(photo.file_id)
        file_name = os.path.join(photo_folder, file_id + ".jpg")
        file_info = bot.get_file(photo.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        with open(file_name, 'wb') as new_file:
            new_file.write(downloaded_file)
        saved_photos.append(file_name)
        bot.reply_to(message, "Фото успешно сохранено!")
        return
    if message.from_user.id in admin_list:
        photo_folder = "temp/изменение_в_расписании"
        saved_photos = []
        photo = message.photo[-1]
        file_id = str(photo.file_id)
        file_name = os.path.join(photo_folder, file_id + ".jpg")
        file_info = bot.get_file(photo.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        with open(file_name, 'wb') as new_file:
            new_file.write(downloaded_file)
        saved_photos.append(file_name)
        bot.reply_to(message, "Фото успешно сохранено!")
        return
    else:
        bot.send_message(message.chat.id,'Я тебя не понял!')

@bot.message_handler(commands=['mem_photo','Mem_photo','мем_photo','Мем_photo',])
def send_mem(message):
    try:
        mems_folder = 'temp/mems/photoes'

        files = os.listdir(mems_folder)

        random_file = random.choice(files)
        bot.send_photo(message.chat.id, photo=open(os.path.join(mems_folder, random_file), 'rb'))
    except:
        bot.send_message(message.chat.id,'Ошибка: некорректный ввод')


@bot.message_handler(commands=['add_user'])
def add_user(message):
    if message.chat.id not in ban_list:
        mas = []
        sqlite_select_query = """SELECT * from students"""
        cursor.execute(sqlite_select_query)
        records = cursor.fetchall()
        for row in records:
            mas.append(row[2])
        if message.chat.id in mas:
            bot.send_message(message.chat.id, 'Вы уже есть в списке')
            return
        else:
            bot.send_message(message.chat.id, 'Введите: Имя')
            bot.register_next_step_handler(message, user_name)
    else:
        bot.send_message(message.chat.id, 'BANNED')


def user_name(msg):
    global memory
    text = msg.text.lower()
    text = text.title()
    memory.append(text)
    bot.send_message(msg.chat.id, 'Введите: Фамилия')
    bot.register_next_step_handler(msg, user_surname)

def user_surname(msg):
    global memory
    text = msg.text.lower()
    text = text.title()
    memory.append(text)
    bot.send_message(msg.chat.id, 'Введите: Класс')
    bot.register_next_step_handler(msg, user_clas)

def user_clas(msg):
    global memory
    text = msg.text.lower()
    text = text.title()
    memory.append(text)
    adduser(msg)

def adduser(message):
    try:
        global memory
        user_id = message.chat.id
        user_name = f'{memory[0]} {memory[1]}'
        user_class = str(memory[2])
        sql_query = "INSERT INTO students (name, tg_id, user_class) VALUES (?, ?, ?)"
        values = (user_name, user_id, user_class)
        cursor.execute(sql_query, values)
        conn.commit()
        memory = []
        bot.send_message(message.chat.id, text=f"Ученик {user_name} добавлен в базу данных.")
    except:
        bot.send_message(message.chat.id,'Ошибка: некорректный ввод')



@bot.message_handler(commands=['add_ticket'])
def add_ticket_handler(message):
    if message.chat.id in ban_list or message.chat.id in standart_list:
        bot.send_message(message.chat.id, 'Ошибка: Недостаточно прав доступа')
    if message.chat.id not in ban_list:
        prompt = message.text.split()
        if len(prompt) == 1:
            bot.send_message(message.chat.id, f'Здравствуй друг! Если вы чем-то недовольны или у вас появилась проблема можете подать заявку в боте или отправить написать модератору.\n\n'\
                    f'почта: <pre>math.lyceum.bot.help@mail.ru</pre>\n', parse_mode='html')
        else:
            user_id = message.chat.id
            problem = message.text.replace('/add_ticket ', '')
            ticket_id = str(user_id) + '_' + str(cursor.lastrowid + 1)
            cursor.execute("INSERT INTO tickets (user_id, problem, ticket_id) VALUES (?, ?, ?)", (user_id, problem, ticket_id))
            conn.commit()
            bot.reply_to(message, f"Ваша проблема добавлена. Номер тикета: {ticket_id}")
    else:
        bot.send_message(message.chat.id,'BANNED')


@bot.message_handler(commands=['tickets'])
def all_print(message):
    if message.chat.id in admin_list:
        sqlite_select_query = """SELECT * from tickets"""
        cursor.execute(sqlite_select_query)
        records = cursor.fetchall()
        bot.send_message(message.chat.id, f'Количество - {len(records)}')
        for row in records:
            mess = f'Тикет номер: <pre>{row[3]}</pre>\n\n' \
                   f'Текст тикета: {str(row[2])}\n\n' \
                   f'Статус: {str(row[4])}'
            bot.send_message(message.chat.id, mess, parse_mode='html')
    else:
        bot.send_message(message.chat.id, 'Ошибка: Недостаточно прав доступа')
        return


@bot.message_handler(commands=['set_ticket_status'])
def set_ticket_status(message):
    if message.chat.id in ban_list or message.chat.id in standart_list:
        bot.send_message(message.chat.id, 'Ошибка: Недостаточно прав доступа')
        return
    else:
        bot.send_message(message.chat.id, 'Отправь мне id тикета')
        bot.register_next_step_handler(message, change_ticket_status)

def change_ticket_status(message):
    global memory2
    memory2.append(message.text)
    button_close = types.InlineKeyboardButton('Закрыть', callback_data='close_ticket')
    button_open = types.InlineKeyboardButton('Открыть', callback_data='open_ticket')
    markup = types.InlineKeyboardMarkup([[button_open,button_close]])
    bot.send_message(message.chat.id, f'Тикет №{message.text}', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "close_ticket")
def close_ticket(call):
    try:
        global memory2
        user_id = memory2[-1].split('_')[0]
        ticket_id = memory2[-1]
        status = 'close'
        if status.lower() not in ('open', 'close', 'открыт', 'закрыт', 'открытый', 'закрытый'):
            bot.send_message(call.message.chat.id, "Неверный статус")
            return
        cursor.execute("UPDATE tickets SET status = ? WHERE ticket_id = ? AND user_id = ?",
                       (status.lower(), ticket_id, user_id))
        conn.commit()
        bot.send_message(call.message.chat.id, f"Статус тикета {ticket_id} изменен на {status.lower()}")
    except:
        bot.send_message(call.message.chat.id, f"Ошибка: Некорректный ввод")


@bot.callback_query_handler(func=lambda call: call.data == "open_ticket")
def close_ticket(call):
    try:
        global memory2
        user_id = memory2[-1].split('_')[0]
        ticket_id = memory2[-1]
        status = 'open'
        if status.lower() not in ('open', 'close', 'открыт', 'закрыт', 'открытый', 'закрытый'):
            bot.send_message(call.message.chat.id, "Неверный статус")
            return
        cursor.execute("UPDATE tickets SET status = ? WHERE ticket_id = ? AND user_id = ?",
                       (status.lower(), ticket_id, user_id))
        conn.commit()
        bot.send_message(call.message.chat.id, f"Статус тикета {ticket_id} изменен на {status.lower()}")
    except:
        bot.send_message(call.message.chat.id, f"Ошибка: Некорректный ввод")


@bot.message_handler(commands=['delete_ticket'])
def delete_ticket(message):
    if message.chat.id not in admin_list:
        bot.send_message(chat_id=message.chat.id, text='У вас нет прав доступа')
        return
    try:
        ticket_id = message.text.split()[1]

        cursor.execute("DELETE FROM tickets WHERE ticket_id=?", (ticket_id,))
        conn.commit()

        bot.send_message(chat_id=message.chat.id, text=f'Тикет {ticket_id} успешно удален')
    except:
        bot.send_message(message.chat.id,'Ошибка: некорректный ввод')



@bot.message_handler(commands=['logs'])
def all_print(message):
    if message.chat.id in admin_list:
        sqlite_select_query = """SELECT * from students"""
        cursor.execute(sqlite_select_query)
        records = cursor.fetchall()
        for row in records:
            mess = f'{row[1]} <pre>{row[2]}</pre> {row[3]} {row[4]}'
            bot.send_message(message.chat.id, mess, parse_mode='html')
    else:
        bot.send_message(message.chat.id, 'Ошибка: Недостаточно прав доступа')


@bot.message_handler(commands=['Site', 'site', 'website', 'сайт', 'Сайт'])
def website(message):
    if message.chat.id not in ban_list:
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("Сайт", url='матлицей.рф'))
        markup.add(types.InlineKeyboardButton("VK-группа", url='https://vk.com/matliceum'))
        markup.add(types.InlineKeyboardButton("Однаклассники", url='https://ok.ru/group/70000001007910'))
        markup.add(types.InlineKeyboardButton("Telegram-группа", url='https://t.me/wxGCj7i8HfE3MDky'))
        bot.send_message(message.chat.id, 'Наши Соц.Сети', reply_markup=markup)
    else:
        bot.send_message(message.chat.id, 'BANNED')


@bot.message_handler(commands=['menu', 'Menu', 'back', 'Назад'])
def help(message):
    if message.chat.id not in ban_list:
        markup_reply = types.ReplyKeyboardMarkup()
        izmenenie = types.KeyboardButton('/Изменение_в_расписании')
        admin = types.KeyboardButton('/admin')
        Sites = types.KeyboardButton('/Сайт')
        menu = types.KeyboardButton('/help')
        classes = types.KeyboardButton('/Классы')
        meme = types.KeyboardButton('/Mem')
        markup_reply.row(izmenenie,admin)
        markup_reply.row(Sites, menu)
        markup_reply.row(classes, meme)
        bot.send_message(message.chat.id, 'Чем могу помочь?', reply_markup=markup_reply)
    else:
        bot.send_message(message.chat.id, 'BANNED')


@bot.message_handler(commands=['Изменение_в_расписании','new_shedule'])
def send_mem(message):
    if message.chat.id not in ban_list:
        photo_dir = 'temp/изменение_в_расписании'
        photos = os.listdir(photo_dir)
        for photo in photos:
            with open(os.path.join(photo_dir, photo), 'rb') as f:
                bot.send_photo(message.chat.id, f)
    else:
        bot.send_message(message.chat.id,'BANNED')

@bot.message_handler(commands=['Класс_10А'])
def cl10A(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 10А')
    Tuesday = types.KeyboardButton('/Вторник 10А')
    Wednesday = types.KeyboardButton('/Среда 10А')
    Thursday = types.KeyboardButton('/Четверг 10А')
    Friday = types.KeyboardButton('/Пятница 10А')
    Saturday = types.KeyboardButton('/Суббота 10А')
    Sunday = types.KeyboardButton('/Воскресенье 10А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_11А'])
def cl11A(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 11А')
    Tuesday = types.KeyboardButton('/Вторник 11А')
    Wednesday = types.KeyboardButton('/Среда 11А')
    Thursday = types.KeyboardButton('/Четверг 11А')
    Friday = types.KeyboardButton('/Пятница 11А')
    Saturday = types.KeyboardButton('/Суббота 11А')
    Sunday = types.KeyboardButton('/Воскресенье 11А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_11Б'])
def cl11B(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 11Б')
    Tuesday = types.KeyboardButton('/Вторник 11Б')
    Wednesday = types.KeyboardButton('/Среда 11Б')
    Thursday = types.KeyboardButton('/Четверг 11Б')
    Friday = types.KeyboardButton('/Пятница 11Б')
    Saturday = types.KeyboardButton('/Суббота 11Б')
    Sunday = types.KeyboardButton('/Воскресенье 11Б')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_9В'])
def cl9C(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 9В')
    Tuesday = types.KeyboardButton('/Вторник 9В')
    Wednesday = types.KeyboardButton('/Среда 9В')
    Thursday = types.KeyboardButton('/Четверг 9В')
    Friday = types.KeyboardButton('/Пятница 9В')
    Saturday = types.KeyboardButton('/Суббота 9В')
    Sunday = types.KeyboardButton('/Воскресенье 9В')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_9Б'])
def cl9B(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 9Б')
    Tuesday = types.KeyboardButton('/Вторник 9Б')
    Wednesday = types.KeyboardButton('/Среда 9Б')
    Thursday = types.KeyboardButton('/Четверг 9Б')
    Friday = types.KeyboardButton('/Пятница 9Б')
    Saturday = types.KeyboardButton('/Суббота 9Б')
    Sunday = types.KeyboardButton('/Воскресенье 9Б')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_9А'])
def cl9A(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 9А')
    Tuesday = types.KeyboardButton('/Вторник 9А')
    Wednesday = types.KeyboardButton('/Среда 9А')
    Thursday = types.KeyboardButton('/Четверг 9А')
    Friday = types.KeyboardButton('/Пятница 9А')
    Saturday = types.KeyboardButton('/Суббота 9А')
    Sunday = types.KeyboardButton('/Воскресенье 9А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_8В'])
def cl8C(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 8В')
    Tuesday = types.KeyboardButton('/Вторник 8В')
    Wednesday = types.KeyboardButton('/Среда 8В')
    Thursday = types.KeyboardButton('/Четверг 8В')
    Friday = types.KeyboardButton('/Пятница 8В')
    Saturday = types.KeyboardButton('/Суббота 8В')
    Sunday = types.KeyboardButton('/Воскресенье 8В')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_8Б'])
def cl8B(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 8Б')
    Tuesday = types.KeyboardButton('/Вторник 8Б')
    Wednesday = types.KeyboardButton('/Среда 8Б')
    Thursday = types.KeyboardButton('/Четверг 8Б')
    Friday = types.KeyboardButton('/Пятница 8Б')
    Saturday = types.KeyboardButton('/Суббота 8Б')
    Sunday = types.KeyboardButton('/Воскресенье 8Б')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_8А'])
def cl8AA(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 8А')
    Tuesday = types.KeyboardButton('/Вторник 8А')
    Wednesday = types.KeyboardButton('/Среда 8А')
    Thursday = types.KeyboardButton('/Четверг 8А')
    Friday = types.KeyboardButton('/Пятница 8А')
    Saturday = types.KeyboardButton('/Суббота 8А')
    Sunday = types.KeyboardButton('/Воскресенье 8А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_7В'])
def cl7C(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 7В')
    Tuesday = types.KeyboardButton('/Вторник 7В')
    Wednesday = types.KeyboardButton('/Среда 7В')
    Thursday = types.KeyboardButton('/Четверг 7В')
    Friday = types.KeyboardButton('/Пятница 7В')
    Saturday = types.KeyboardButton('/Суббота 7В')
    Sunday = types.KeyboardButton('/Воскресенье 7В')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_7Б'])
def cl7B(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 7Б')
    Tuesday = types.KeyboardButton('/Вторник 7Б')
    Wednesday = types.KeyboardButton('/Среда 7Б')
    Thursday = types.KeyboardButton('/Четверг 7Б')
    Friday = types.KeyboardButton('/Пятница 7Б')
    Saturday = types.KeyboardButton('/Суббота 7Б')
    Sunday = types.KeyboardButton('/Воскресенье 7Б')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_7А'])
def cl7A(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 7А')
    Tuesday = types.KeyboardButton('/Вторник 7А')
    Wednesday = types.KeyboardButton('/Среда 7А')
    Thursday = types.KeyboardButton('/Четверг 7А')
    Friday = types.KeyboardButton('/Пятница 7А')
    Saturday = types.KeyboardButton('/Суббота 7А')
    Sunday = types.KeyboardButton('/Воскресенье 7А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_6Б'])
def cl6B(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 6Б')
    Tuesday = types.KeyboardButton('/Вторник 6Б')
    Wednesday = types.KeyboardButton('/Среда 6Б')
    Thursday = types.KeyboardButton('/Четверг 6Б')
    Friday = types.KeyboardButton('/Пятница 6Б')
    Saturday = types.KeyboardButton('/Суббота 6Б')
    Sunday = types.KeyboardButton('/Воскресенье 6Б')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_6А'])
def cl6A(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 6А')
    Tuesday = types.KeyboardButton('/Вторник 6А')
    Wednesday = types.KeyboardButton('/Среда 6А')
    Thursday = types.KeyboardButton('/Четверг 6А')
    Friday = types.KeyboardButton('/Пятница 6А')
    Saturday = types.KeyboardButton('/Суббота 6А')
    Sunday = types.KeyboardButton('/Воскресенье 6А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_5В'])
def cl5C(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 5В')
    Tuesday = types.KeyboardButton('/Вторник 5В')
    Wednesday = types.KeyboardButton('/Среда 5В')
    Thursday = types.KeyboardButton('/Четверг 5В')
    Friday = types.KeyboardButton('/Пятница 5В')
    Saturday = types.KeyboardButton('/Суббота 5В')
    Sunday = types.KeyboardButton('/Воскресенье 5В')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_5Б'])
def cl5B(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 5Б')
    Tuesday = types.KeyboardButton('/Вторник 5Б')
    Wednesday = types.KeyboardButton('/Среда 5Б')
    Thursday = types.KeyboardButton('/Четверг 5Б')
    Friday = types.KeyboardButton('/Пятница 5Б')
    Saturday = types.KeyboardButton('/Суббота 5Б')
    Sunday = types.KeyboardButton('/Воскресенье 5Б')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Класс_5А'])
def cl5A(message):
    markup_classes = types.ReplyKeyboardMarkup()
    Monday = types.KeyboardButton('/Понедельник 5А')
    Tuesday = types.KeyboardButton('/Вторник 5А')
    Wednesday = types.KeyboardButton('/Среда 5А')
    Thursday = types.KeyboardButton('/Четверг 5А')
    Friday = types.KeyboardButton('/Пятница 5А')
    Saturday = types.KeyboardButton('/Суббота 5А')
    Sunday = types.KeyboardButton('/Воскресенье 5А')
    markup_classes.row(Monday, Tuesday, Wednesday)
    markup_classes.row(Thursday, Friday, Saturday)
    markup_classes.row(Sunday)
    Back = types.KeyboardButton('/Назад')
    markup_classes.row(Back)
    bot.send_message(message.chat.id, 'Какой день недели нужен?', reply_markup=markup_classes)


@bot.message_handler(commands=['Понедельник'])
def ponedelnik(message):
    Day = 1
    params = message.text.split()
    school_class = params[1]
    if school_class == '10А':
        filename = '10а.xls'
    elif school_class == '11А':
        filename = '11а.xls'
    elif school_class == '11Б':
        filename = '11б.xls'
    elif school_class == '9А':
        filename = '9а.xls'
    elif school_class == '9Б':
        filename = '9б.xls'
    elif school_class == '9В':
        filename = '9в.xls'
    elif school_class == '8А':
        filename = '8а.xls'
    elif school_class == '8Б':
        filename = '8б.xls'
    elif school_class == '8В':
        filename = '8в.xls'
    elif school_class == '7А':
        filename = '7а.xls'
    elif school_class == '7Б':
        filename = '7б.xls'
    elif school_class == '7В':
        filename = '7в.xls'
    elif school_class == '6А':
        filename = '6а.xls'
    elif school_class == '6Б':
        filename = '6б.xls'
    elif school_class == '5А':
        filename = '5а.xls'
    elif school_class == '5Б':
        filename = '5б.xls'
    elif school_class == '5В':
        filename = '5в.xls'
    with open(os.path.join(os.getcwd(), filename), 'r') as f:  # open in readonly mode
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        for i in range(Day - 1, Day):
            result = []
            for j in range(0, 6):
                string = ''
                if sheet.cell_value(i, j) != 'None':
                    string = f'{j + 1}) {sheet.cell_value(i, j)}\n\n'
                    result.append(string)
            for i in result:
                bot.send_message(message.chat.id, i)


@bot.message_handler(commands=['Вторник'])
def vtornik(message):
    Day = 2
    params = message.text.split()
    school_class = params[1]
    if school_class == '10А':
        filename = '10а.xls'
    elif school_class == '11А':
        filename = '11а.xls'
    elif school_class == '11Б':
        filename = '11б.xls'
    elif school_class == '9А':
        filename = '9а.xls'
    elif school_class == '9Б':
        filename = '9б.xls'
    elif school_class == '9В':
        filename = '9в.xls'
    elif school_class == '8А':
        filename = '8а.xls'
    elif school_class == '8Б':
        filename = '8б.xls'
    elif school_class == '8В':
        filename = '8в.xls'
    elif school_class == '7А':
        filename = '7а.xls'
    elif school_class == '7Б':
        filename = '7б.xls'
    elif school_class == '7В':
        filename = '7в.xls'
    elif school_class == '6А':
        filename = '6а.xls'
    elif school_class == '6Б':
        filename = '6б.xls'
    elif school_class == '5А':
        filename = '5а.xls'
    elif school_class == '5Б':
        filename = '5б.xls'
    elif school_class == '5В':
        filename = '5в.xls'
    with open(os.path.join(os.getcwd(), filename), 'r') as f:  # open in readonly mode
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        for i in range(Day - 1, Day):
            result = []
            for j in range(0, 6):
                string = ''
                if sheet.cell_value(i, j) != 'None':
                    string = f'{j + 1}) {sheet.cell_value(i, j)}\n\n'
                    result.append(string)
            for i in result:
                bot.send_message(message.chat.id, i)


@bot.message_handler(commands=['Среда'])
def sreda(message):
    Day = 3
    params = message.text.split()
    school_class = params[1]
    if school_class == '10А':
        filename = '10а.xls'
    elif school_class == '11А':
        filename = '11а.xls'
    elif school_class == '11Б':
        filename = '11б.xls'
    elif school_class == '9А':
        filename = '9а.xls'
    elif school_class == '9Б':
        filename = '9б.xls'
    elif school_class == '9В':
        filename = '9в.xls'
    elif school_class == '8А':
        filename = '8а.xls'
    elif school_class == '8Б':
        filename = '8б.xls'
    elif school_class == '8В':
        filename = '8в.xls'
    elif school_class == '7А':
        filename = '7а.xls'
    elif school_class == '7Б':
        filename = '7б.xls'
    elif school_class == '7В':
        filename = '7в.xls'
    elif school_class == '6А':
        filename = '6а.xls'
    elif school_class == '6Б':
        filename = '6б.xls'
    elif school_class == '5А':
        filename = '5а.xls'
    elif school_class == '5Б':
        filename = '5б.xls'
    elif school_class == '5В':
        filename = '5в.xls'
    with open(os.path.join(os.getcwd(), filename), 'r') as f:  # open in readonly mode
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        for i in range(Day - 1, Day):
            result = []
            for j in range(0, 6):
                string = ''
                if sheet.cell_value(i, j) != 'None':
                    string = f'{j + 1}) {sheet.cell_value(i, j)}\n\n'
                    result.append(string)
            for i in result:
                bot.send_message(message.chat.id, i)
                # print(*result)


@bot.message_handler(commands=['Четверг'])
def chetverg(message):
    Day = 4
    params = message.text.split()
    school_class = params[1]
    if school_class == '10А':
        filename = '10а.xls'
    elif school_class == '11А':
        filename = '11а.xls'
    elif school_class == '11Б':
        filename = '11б.xls'
    elif school_class == '9А':
        filename = '9а.xls'
    elif school_class == '9Б':
        filename = '9б.xls'
    elif school_class == '9В':
        filename = '9в.xls'
    elif school_class == '8А':
        filename = '8а.xls'
    elif school_class == '8Б':
        filename = '8б.xls'
    elif school_class == '8В':
        filename = '8в.xls'
    elif school_class == '7А':
        filename = '7а.xls'
    elif school_class == '7Б':
        filename = '7б.xls'
    elif school_class == '7В':
        filename = '7в.xls'
    elif school_class == '6А':
        filename = '6а.xls'
    elif school_class == '6Б':
        filename = '6б.xls'
    elif school_class == '5А':
        filename = '5а.xls'
    elif school_class == '5Б':
        filename = '5б.xls'
    elif school_class == '5В':
        filename = '5в.xls'
    with open(os.path.join(os.getcwd(), filename), 'r') as f:  # open in readonly mode
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        for i in range(Day - 1, Day):
            result = []
            for j in range(0, 6):
                string = ''
                if sheet.cell_value(i, j) != 'None':
                    string = f'{j + 1}) {sheet.cell_value(i, j)}\n\n'
                    result.append(string)
            for i in result:
                bot.send_message(message.chat.id, i)


@bot.message_handler(commands=['Пятница'])
def pyatniza(message):
    Day = 5
    params = message.text.split()
    school_class = params[1]
    if school_class == '10А':
        filename = '10а.xls'
    elif school_class == '11А':
        filename = '11а.xls'
    elif school_class == '11Б':
        filename = '11б.xls'
    elif school_class == '9А':
        filename = '9а.xls'
    elif school_class == '9Б':
        filename = '9б.xls'
    elif school_class == '9В':
        filename = '9в.xls'
    elif school_class == '8А':
        filename = '8а.xls'
    elif school_class == '8Б':
        filename = '8б.xls'
    elif school_class == '8В':
        filename = '8в.xls'
    elif school_class == '7А':
        filename = '7а.xls'
    elif school_class == '7Б':
        filename = '7б.xls'
    elif school_class == '7В':
        filename = '7в.xls'
    elif school_class == '6А':
        filename = '6а.xls'
    elif school_class == '6Б':
        filename = '6б.xls'
    elif school_class == '5А':
        filename = '5а.xls'
    elif school_class == '5Б':
        filename = '5б.xls'
    elif school_class == '5В':
        filename = '5в.xls'
    with open(os.path.join(os.getcwd(), filename), 'r') as f:  # open in readonly mode
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        for i in range(Day - 1, Day):
            result = []
            for j in range(0, 6):
                string = ''
                if sheet.cell_value(i, j) != 'None':
                    string = f'{j + 1}) {sheet.cell_value(i, j)}\n\n'
                    result.append(string)
            for i in result:
                bot.send_message(message.chat.id, i)


@bot.message_handler(commands=['Суббота'])
def subbota(message):
    Day = 6
    params = message.text.split()
    school_class = params[1]
    if school_class == '10А':
        filename = '10а.xls'
    elif school_class == '11А':
        filename = '11а.xls'
    elif school_class == '11Б':
        filename = '11б.xls'
    elif school_class == '9А':
        filename = '9а.xls'
    elif school_class == '9Б':
        filename = '9б.xls'
    elif school_class == '9В':
        filename = '9в.xls'
    elif school_class == '8А':
        filename = '8а.xls'
    elif school_class == '8Б':
        filename = '8б.xls'
    elif school_class == '8В':
        filename = '8в.xls'
    elif school_class == '7А':
        filename = '7а.xls'
    elif school_class == '7Б':
        filename = '7б.xls'
    elif school_class == '7В':
        filename = '7в.xls'
    elif school_class == '6А':
        filename = '6а.xls'
    elif school_class == '6Б':
        filename = '6б.xls'
    elif school_class == '5А':
        filename = '5а.xls'
    elif school_class == '5Б':
        filename = '5б.xls'
    elif school_class == '5В':
        filename = '5в.xls'
    with open(os.path.join(os.getcwd(), filename), 'r') as f:  # open in readonly mode
        book = xlrd.open_workbook(filename)
        sheet = book.sheet_by_index(0)
        for i in range(Day - 1, Day):
            result = []
            for j in range(0, 6):
                string = ''
                if sheet.cell_value(i, j) != 'None':
                    string = f'{j + 1}) {sheet.cell_value(i, j)}\n\n'
                    result.append(string)
            for i in result:
                bot.send_message(message.chat.id, i)


@bot.message_handler(commands=['Mem','mem','мем','Мем'])
def knopk_mema(message):
    if message.chat.id not in ban_list:
        markup_mems = types.ReplyKeyboardMarkup()
        meme_text = types.KeyboardButton('/Mem_text')
        meme_photo = types.KeyboardButton('/Mem_photo')
        markup_mems.row(meme_text,meme_photo)
        Back = types.KeyboardButton('/Назад')
        markup_mems.row(Back)
        bot.send_message(message.chat.id, 'Какой мем?', reply_markup=markup_mems)
    else:
        bot.send_message(message.chat.id,'BANNED')


@bot.message_handler(commands=['Классы', 'classes'])
def raspisanie_knopki(message):
    if message.chat.id not in ban_list:
        markup_classes = types.ReplyKeyboardMarkup()
        class_10A = types.KeyboardButton('/Класс_10А')
        class_11A = types.KeyboardButton('/Класс_11А')
        class_11B = types.KeyboardButton('/Класс_11Б')
        class_9B = types.KeyboardButton('/Класс_9Б')
        class_9A = types.KeyboardButton('/Класс_9А')
        class_9C = types.KeyboardButton('/Класс_9В')
        class_8A = types.KeyboardButton('/Класс_8А')
        class_8B = types.KeyboardButton('/Класс_8Б')
        class_8C = types.KeyboardButton('/Класс_8В')
        class_7A = types.KeyboardButton('/Класс_7А')
        class_7B = types.KeyboardButton('/Класс_7Б')
        class_7C = types.KeyboardButton('/Класс_7В')
        class_5A = types.KeyboardButton('/Класс_5А')
        class_5B = types.KeyboardButton('/Класс_5Б')
        class_5C = types.KeyboardButton('/Класс_5В')
        class_6A = types.KeyboardButton('/Класс_6А')
        class_6B = types.KeyboardButton('/Класс_6Б')
        markup_classes.row(class_11B, class_11A, class_10A)
        markup_classes.row(class_9A, class_9B, class_9C)
        markup_classes.row(class_8A, class_8B, class_8C)
        markup_classes.row(class_7A, class_7B, class_7C)
        markup_classes.row(class_6A, class_6B)
        markup_classes.row(class_5A, class_5B, class_5C)
        Back = types.KeyboardButton('/Назад')
        markup_classes.row(Back)
        bot.send_message(message.chat.id, 'Какой тебе нужный класс?', reply_markup=markup_classes)
    else:
        bot.send_message(message.chat.id, 'BANNED')


@bot.message_handler(commands=['Воскресенье'])
def vosk(message):
    b = randint(1, 3)
    if b == 1:
        mem1 = open("1.jpg", "rb")
        bot.send_photo(message.chat.id, mem1)
    if b == 2:
        mem2 = open("2.jpg", "rb")
        bot.send_photo(message.chat.id, mem2)
    if b == 3:
        mem3 = open("3.jpg", "rb")
        bot.send_photo(message.chat.id, mem3)
    bot.send_message(message.chat.id, 'Выспись в воскресенье!', parse_mode='html')


@bot.message_handler(commands=['Мем_text', 'Mem_text', 'mem_text','мем_text'])
def mem(message):
    if message.chat.id not in ban_list:
        # bot.send_message(message.chat.id, mems.MEM)
        bot.send_message(message.chat.id, 'ОШИБКА: Технические неполадки')
    else:
        bot.send_message(message.chat.id, 'BANNED')


@bot.message_handler(commands=['Id', 'id'])
def id(message):
    if message.chat.id not in ban_list:
        mess = f'Ваш ID:\n' \
               f'<pre>{message.chat.id}</pre>'
        bot.send_message(message.chat.id, mess, parse_mode='html')
    else:
        bot.send_message(message.chat.id, 'BANNED')

def find_user_by_name(name):
    spisoc = []
    if name in users_dict:
        spisoc.append(users_dict[name]['name'])
        return spisoc
    else:
        return None

def get_tg_id_by_name(name):
    if name in users_dict:
        return users_dict[name]['tg_id']
    else:
        return None

@bot.message_handler(commands=['find'])
def handle_find_command(message):
    # Отправляем сообщение с просьбой ввести имя пользователя
    bot.reply_to(message, "Введите имя пользователя:")
    # Устанавливаем состояние пользователя в "ожидание имени"
    bot.register_next_step_handler(message, handle_name_input)

def handle_name_input(msg):
    flag = False
    name = msg.text
    name = name.lower()
    name = name.title()
    # Получаем введенное имя пользователя
    if len(msg.text.split()) == 2:
        all_name = msg.text.lower()
        name1 = all_name.split()[0]
        name1 = name1.title()
        name2 = all_name.split()[1]
        name2 = name2.title()
        name_reverse = f'{name2} {name1}'
        for key in users_dict:
            key1 = users_dict[key]['name'].split()[0]
            key2 = users_dict[key]['name'].split()[1]
            if name1 in [key1, key2] or name2 in [key1, key2]:
                tg_id1 = get_tg_id_by_name(name)
                tg_id2 = get_tg_id_by_name(name_reverse)
                if tg_id1 is not None:
                    bot.reply_to(msg, f"Пользователь '{name}' <pre>{tg_id1}</pre>", parse_mode='html')
                elif tg_id2 is not None:
                    bot.reply_to(msg, f"Пользователь '{name_reverse}' <pre>{tg_id2}</pre>", parse_mode='html')
                else:
                    bot.reply_to(msg, f"Пользователь '{name}' не найден", parse_mode='html')

    if len(msg.text.split()) == 1:
        for key in users_dict:
            key1 = users_dict[key]['name'].split()[0]
            key2 = users_dict[key]['name'].split()[1]
            if name in [key1, key2]:
                all_name1 = f'{key1} {key2}'
                all_name2 = f'{key2} {key1}'
                tg_id1 = get_tg_id_by_name(all_name1)
                tg_id2 = get_tg_id_by_name(all_name2)
                if tg_id1 is not None:
                    bot.reply_to(msg, f"Пользователь '{all_name1}' <pre>{tg_id1}</pre>", parse_mode='html')
                    flag = True
                elif tg_id2 is not None:
                    flag = True
                    bot.reply_to(msg, f"Пользователь '{all_name2}' <pre>{tg_id2}</pre>", parse_mode='html')
            else:
                if flag == True:
                    flag = True
                else:
                    flag = False
                # bot.reply_to(msg, f"Пользователь '{name}' не найден", parse_mode='html')

    if flag == False:
        bot.reply_to(msg, f"Пользователь '{msg.text}' не найден", parse_mode='html')

@bot.message_handler(commands=['send_text'])
def letter(message):
    if message.chat.id in admin_list:
        args = message.text.split()[1]
        if int(args) not in ban_list and int(args) not in admin_list and int(args) not in standart_list and int(args) not in moder_list:
            bot.send_message(message.chat.id, "Неверный формат команды")
            # for i in args:
            bot.send_message(message.chat.id,args)
            #     bot.send_message(message.chat.id,i)
            return
        user_id = int(args)
        send_text = message.text.replace(f'/send_text {user_id}','')
        bot.send_message(user_id,send_text,parse_mode='html')
        bot.send_message(message.chat.id, f"Я отправил ваше сообщение пользователю {user_id}")
    else:
        bot.send_message(message.chat.id, "Недостаточно прав доступа")


@bot.callback_query_handler(func=lambda call: call.data == 'message4all')
def rassilki(call):
    button_delete_message = types.InlineKeyboardButton('Удалить рассылку', callback_data='delete_MFA')
    button_new_message = types.InlineKeyboardButton('Добавить рассылку', callback_data='new_MFA')
    button_logs_message = types.InlineKeyboardButton('Все рассылки', callback_data='logs_MFA')
    button_back = types.InlineKeyboardButton('Назад', callback_data='command_admin_room')
    markup = types.InlineKeyboardMarkup([[button_delete_message, button_new_message], [button_back,button_logs_message]])
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Кабинет администратора', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'logs_MFA')
def logs_MFA(call):
    filename = 'рассылка.xlsx'
    mas = []
    stroka = []
    wb = load_workbook(filename)
    ws = wb.active
    mess = ''
    for row in ws.iter_rows(min_row=2, max_col=8, max_row=100, values_only=True):
        for cell in row:
            stroka.append(cell)
        if stroka[0] != None:
            mas.append(stroka)
        stroka = []
    for lists in mas:
        for cell in lists:
            mess += f' {str(cell)}'
        mess += '\n\n'
        bot.send_message(chat_id=call.message.chat.id, text=mess)
        mess = ''


# Обработчик нажатия кнопки
@bot.callback_query_handler(func=lambda call: call.data == "command_admin_room")
def command_handler(call):
    button_delete = types.InlineKeyboardButton('Очистить содержимое', callback_data='clear_izmenenie')
    button_new_admin = types.InlineKeyboardButton('Добавить администратора', callback_data='user_to_admin_room')
    button_MFA = types.InlineKeyboardButton('Рассылки', callback_data='message4all')
    button_tickets = types.InlineKeyboardButton('Тикеты', callback_data='tickets')
    button_back = types.InlineKeyboardButton('Назад', callback_data='go2admin_menu')
    markup = types.InlineKeyboardMarkup([[button_delete, button_new_admin, button_tickets], [button_back, button_MFA]])
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Кабинет администратора', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "go2admin_menu")
def admin_menu(call):
    button = types.InlineKeyboardButton("Управление", callback_data="command_admin_room")
    # Создаем разметку с кнопкой
    markup = types.InlineKeyboardMarkup([[button]])
    # Отправляем сообщение с разметкой
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text="Зайти в кабинет", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "user_to_admin_room")
def user_to_admin_room(call):
    button1 = types.InlineKeyboardButton('Добавить администратора', callback_data='add_admin')
    button2 = types.InlineKeyboardButton('Удалить администратора', callback_data='delete_admin')
    back = types.InlineKeyboardButton('Назад', callback_data='command_admin_room')
    markup = types.InlineKeyboardMarkup([[button1,button2], [back]])
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Вы вошли в режим изменения статуса', reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == "add_admin")
def add_admin(call):
    bot.register_next_step_handler(call.message, add_new_admin)
    bot.send_message(call.message.chat.id, 'Отправь мне id нового администратора')

def add_new_admin(message):
    try:
        global standart_list, admin_list

        cursor.execute("UPDATE students SET role = ? WHERE tg_id = ?", ('admin', int(message.text)))
        conn.commit()

        cursor.execute("SELECT tg_id FROM students WHERE role='standart'")
        results = cursor.fetchall()
        standart_list = []
        for row in results:
            standart_list.append(row[0])

        cursor.execute("SELECT tg_id FROM students WHERE role='admin'")
        results = cursor.fetchall()
        admin_list = []
        for row in results:
            admin_list.append(row[0])

        bot.send_message(message.chat.id, 'Поменял!')
    except:
        bot.send_message(message.chat.id, 'Ошибка: Некорректный id')


@bot.callback_query_handler(func=lambda call: call.data == "delete_admin")
def add_admin(call):
    bot.register_next_step_handler(call.message, delete_old_admin)
    bot.send_message(call.message.chat.id, 'Отправь мне id администратора')


def delete_old_admin(message):
    global standart_list, admin_list
    try:
        cursor.execute("UPDATE students SET role = ? WHERE tg_id = ?", ('standart', int(message.text)))
        conn.commit()

        cursor.execute("SELECT tg_id FROM students WHERE role='standart'")
        standart_list = []
        results = cursor.fetchall()
        for row in results:
            standart_list.append(row[0])

        cursor.execute("SELECT tg_id FROM students WHERE role='admin'")
        standart_list = []
        results = cursor.fetchall()
        for row in results:
            admin_list.append(row[0])

        bot.send_message(message.chat.id, 'Поменял!')
    except:
        bot.send_message(message.chat.id, 'Ошибка: Некорректный id')


@bot.callback_query_handler(func=lambda call: call.data == 'clear_izmenenie')
def delete_photoes(call):
    try:
        path = 'temp/изменение_в_расписании'
        for file_name in os.listdir(path):
            os.remove(os.path.join(path, file_name))
        bot.send_message(chat_id=call.message.chat.id, text="Фотографии успешно удалены!")
    except IndexError:
        bot.reply_to(chat_id=call.message.chat.id, message=call.message.message_id, text="Вы не указали путь к директории!")


# Обработчик команды
@bot.message_handler(commands=['Admin', 'admin'])
def admin_handler(message):
    if message.chat.id not in admin_list:
        bot.send_message(message.chat.id, 'Ошибка: Недостаточно прав доступа')
        return
    # Создаем кнопку
    button = types.InlineKeyboardButton("Управление", callback_data="command_admin_room")
    # Создаем разметку с кнопкой
    markup = types.InlineKeyboardMarkup([[button]])
    # Отправляем сообщение с разметкой
    bot.send_message(message.chat.id, "Зайти в кабинет", reply_markup=markup)


@bot.callback_query_handler(func=lambda call: call.data == 'tickets')
def tickets_room(call):
    button_tickets = types.InlineKeyboardButton('Все тикеты', callback_data='all_tickets')
    button_delete_ticket = types.InlineKeyboardButton('Удалить', callback_data='delete_ticket')
    button_back = types.InlineKeyboardButton('Назад', callback_data='command_admin_room')
    markup = types.InlineKeyboardMarkup([[button_tickets, button_delete_ticket], [button_back]])
    bot.edit_message_text(chat_id=call.message.chat.id, message_id=call.message.message_id, text='Комната тикетов', reply_markup=markup)

@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    if call.data == 'all_tickets':
        all_print1(call.message)
    elif call.data == 'delete_ticket':
        delete_ticket1(call.message)
    elif call.data == 'command_admin_room':
        command_handler(call.message)

def all_print1(message):
    if message.chat.id in admin_list:
        sqlite_select_query = """SELECT * from tickets"""
        cursor.execute(sqlite_select_query)
        records = cursor.fetchall()
        bot.send_message(message.chat.id, f'Количество - {len(records)}')
        for row in records:
            mess = f'Тикет номер: <pre>{str(row[3])}</pre>\n\n' \
                   f'Текст тикета: {str(row[2])}\n\n' \
                   f'Статус: {str(row[4])}'
            bot.send_message(message.chat.id, mess, parse_mode='html')
    else:
        bot.send_message(message.chat.id, 'Ошибка: Недостаточно прав доступа')
        return

def delete_ticket1(message):
    if message.chat.id not in admin_list:
        bot.send_message(chat_id=message.chat.id, text='У вас нет прав доступа')
        return
    try:
        bot.send_message(message.chat.id, 'Отправь мне id тикета')
        bot.register_next_step_handler(message, delete_ticket2)
    except:
        bot.send_message(message.chat.id, 'Ошибка: некорректный ввод')

def delete_ticket2(message):
    try:
        ticket_id = message.text

        cursor.execute("DELETE FROM tickets WHERE ticket_id=?", (ticket_id,))
        conn.commit()

        bot.send_message(chat_id=message.chat.id, text=f'Тикет {ticket_id} успешно удален')
    except:
        bot.send_message(message.chat.id,'Ошибка: некорректный ввод')

@bot.message_handler(content_types=['text'])
def get_user_text(message):
    if message.chat.id not in ban_list:
        text = message.text
        if text.endswith('!' or '.'):
            text = text[:-1]
        if text in ["Привет", "привет", "Hi", 'HI', "Hello", 'hi']:
            text = text + ' 🖐🏻'
            bot.send_message(message.chat.id, text, parse_mode='html')
            return
        if text.lower() == 'спасибо':
            bot.send_message(message.chat.id, 'Пожалуйста! ❤️😘')
            return
        elif message.text in ["Id", 'id', 'ID', 'id']:
            bot.send_message(message.chat.id, f"Твой ID: {message.chat.id}", parse_mode='html')
            return
        else:
            bot.send_message(message.chat.id, 'Я тебя не понимаю')
    else:
        bot.send_message(message.chat.id, 'BANNED')


@bot.message_handler(content_types=['sticker'])
def sticker(msg):
    if msg.chat.id not in ban_list:
        bot.send_sticker(msg.chat.id, 'CAACAgIAAxkBAAIo6WQiv8t8A6YnidBbBmMlAZSSOx24AAJUAANBtVYMarf4xwiNAfovBA')
    else:
        bot.send_message(msg.chat.id, 'BANNED')
bot.infinity_polling()
